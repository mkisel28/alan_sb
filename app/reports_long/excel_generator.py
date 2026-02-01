"""
Генератор Excel отчетов
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict
from .utils import get_platform_name


class ExcelReportGenerator:
    """Генерирует отчеты в формате Excel"""

    def __init__(self):
        self.wb = None

    def generate(self, data: Dict) -> BytesIO:
        """
        Генерирует Excel отчет

        Args:
            data: Данные аналитики из API

        Returns:
            BytesIO с Excel файлом
        """
        self.wb = Workbook()

        # Удаляем дефолтный лист
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Создаем сводный лист
        self._create_summary_sheet(data)

        # Создаем листы по платформам
        platforms = data.get("platforms", {})
        for platform_key, platform_data in platforms.items():
            self._create_platform_sheet(
                platform_key, platform_data, data.get("previous_period")
            )

        # Сохраняем в BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        return output

    def _create_summary_sheet(self, data: Dict):
        """Создает сводный лист"""
        ws = self.wb.create_sheet("Сводка")

        # Заголовок
        ws["A1"] = "СВОДНЫЙ ОТЧЕТ"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 30

        # Период
        period = data.get("period", {})
        ws["A3"] = "Период анализа:"
        ws["B3"] = f"{period.get('start', '')[:10]} — {period.get('end', '')[:10]}"
        ws["A3"].font = Font(bold=True)

        # Заголовки таблицы
        headers = [
            "Платформа",
            "Авторов",
            "Подписчики",
            "Посты",
            "Просмотры",
            "Вовлечения",
            "Avg PS",
            "Avg ER",
        ]

        row = 5
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Данные
        platforms = data.get("platforms", {})
        row = 6

        for platform_key, platform_data in platforms.items():
            agg = platform_data["aggregated"]
            platform_name = get_platform_name(platform_key)

            ws.cell(row=row, column=1, value=platform_name)
            ws.cell(row=row, column=2, value=agg["total_authors"])
            ws.cell(row=row, column=3, value=agg["total_followers"])
            ws.cell(row=row, column=4, value=agg["total_posts"])
            ws.cell(row=row, column=5, value=agg["total_views"])
            ws.cell(row=row, column=6, value=agg["total_engagement"])
            ws.cell(row=row, column=7, value=round(agg["avg_PS"], 2))
            ws.cell(row=row, column=8, value=round(agg["avg_ER_view"] * 100, 2))

            row += 1

        # Автоширина колонок
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Границы
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_cells in ws[f"A5:{get_column_letter(8)}{row - 1}"]:
            for cell in row_cells:
                cell.border = thin_border

    def _create_platform_sheet(
        self, platform_key: str, platform_data: Dict, prev_period: Dict = None
    ):
        """Создает лист по платформе"""
        platform_name = get_platform_name(platform_key)
        ws = self.wb.create_sheet(platform_name)

        # Заголовок
        ws["A1"] = platform_name
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:P1")
        ws.row_dimensions[1].height = 25

        # Заголовки основной таблицы
        has_previous = prev_period is not None

        headers = [
            "Автор",
            "Username",
            "Подписчики",
            "Δ Подписчики",
            "Посты",
            "Просмотры",
            "Средние просмотры",
            "Вовлечения",
            "ER_view (%)",
            "SR (%)",
            "CR (%)",
            "PS",
            "MS",
        ]

        if has_previous:
            headers.extend(["Пред. PS", "Δ PS"])

        row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Данные авторов
        row = 4
        authors = platform_data["authors"]

        for author_data in authors:
            metrics = author_data["metrics"]
            scores = author_data["scores"]

            col = 1
            ws.cell(row=row, column=col, value=author_data["author_name"])
            col += 1
            ws.cell(row=row, column=col, value=author_data["username"])
            col += 1
            ws.cell(row=row, column=col, value=metrics["F"])
            col += 1
            ws.cell(row=row, column=col, value=metrics["delta_F"])
            col += 1
            ws.cell(row=row, column=col, value=metrics["P"])
            col += 1
            ws.cell(row=row, column=col, value=metrics["V"])
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["V_avg"], 2))
            col += 1
            ws.cell(row=row, column=col, value=metrics["E"])
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["ER_view"] * 100, 2))
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["SR"] * 100, 4))
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["CR"] * 100, 4))
            col += 1
            ws.cell(row=row, column=col, value=scores["PS"])
            col += 1
            ws.cell(row=row, column=col, value=scores.get("MS", ""))
            col += 1

            if has_previous:
                ws.cell(row=row, column=col, value=scores.get("prev_PS", ""))
                col += 1
                if "prev_PS" in scores:
                    delta_ps = scores["PS"] - scores["prev_PS"]
                    ws.cell(row=row, column=col, value=round(delta_ps, 2))
                col += 1

            row += 1

        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            if col <= 2:
                ws.column_dimensions[get_column_letter(col)].width = 18
            else:
                ws.column_dimensions[get_column_letter(col)].width = 13

        # Границы
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_cells in ws[f"A3:{get_column_letter(len(headers))}{row - 1}"]:
            for cell in row_cells:
                cell.border = thin_border

        # Форматирование чисел
        for r in range(4, row):
            for c in [3, 4, 5, 6, 7, 8]:  # Числовые колонки
                cell = ws.cell(row=r, column=c)
                cell.number_format = "#,##0"

            for c in [9, 10, 11]:  # Проценты
                cell = ws.cell(row=r, column=c)
                cell.number_format = "0.00"

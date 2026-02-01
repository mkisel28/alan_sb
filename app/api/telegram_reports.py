"""
API для генерации отчетов по Telegram каналам
"""

from datetime import datetime
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from api.telegram_analytics import (
    _calculate_telegram_period_metrics,
    _calculate_comparison_metrics,
)
from models import SocialAccount
from reports_telegram.excel_generator import generate_telegram_excel_report

router = APIRouter(prefix="/api/telegram-reports", tags=["telegram-reports"])


@router.get("/excel/all-authors")
async def generate_all_authors_telegram_excel(
    current_start: datetime = Query(..., description="Начало текущего периода"),
    current_end: datetime = Query(..., description="Конец текущего периода"),
    previous_start: datetime | None = Query(
        None, description="Начало прошлого периода"
    ),
    previous_end: datetime | None = Query(None, description="Конец прошлого периода"),
):
    """
    Генерирует Excel отчет для всех авторов с Telegram каналами

    Возвращает файл Excel с аналитикой по всем авторам
    """
    from models import Author
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Получаем данные всех авторов
    authors = await Author.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Все авторы Telegram"

    # Заголовки
    headers = [
        "Автор",
        "Каналов",
        "Подписчики",
        "Публикации",
        "Просмотры",
        "Ср. просмотры",
        "Вовлечения",
        "Ср. вовлечения",
        "ER view %",
        "ERR %",
        "ER %",
        "Ср. охват поста",
        "ИЦ",
        "Реакции",
        "Комментарии",
        "Пересылки",
    ]

    # Стиль заголовков
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Данные
    row = 2
    for author in authors:
        telegram_accounts = await SocialAccount.filter(
            author_id=author.id, platform="telegram", is_active=True
        ).all()

        if not telegram_accounts:
            continue

        # Собираем метрики
        aggregated = {
            "F": 0,
            "P": 0,
            "V": 0,
            "E": 0,
            "total_reactions": 0,
            "total_comments": 0,
            "total_shares": 0,
            "err_percent_sum": 0,
            "er_percent_sum": 0,
            "ci_index_sum": 0,
            "avg_post_reach_sum": 0,
        }

        for account in telegram_accounts:
            metrics = await _calculate_telegram_period_metrics(
                account, current_start, current_end
            )

            aggregated["F"] += metrics.get("F", 0)
            aggregated["P"] += metrics.get("P", 0)
            aggregated["V"] += metrics.get("V", 0)
            aggregated["E"] += metrics.get("E", 0)
            aggregated["total_reactions"] += metrics.get("total_reactions", 0)
            aggregated["total_comments"] += metrics.get("total_comments", 0)
            aggregated["total_shares"] += metrics.get("total_shares", 0)
            aggregated["err_percent_sum"] += metrics.get("ERR_percent", 0)
            aggregated["er_percent_sum"] += metrics.get("ER_percent", 0)
            aggregated["ci_index_sum"] += metrics.get("ci_index", 0)
            aggregated["avg_post_reach_sum"] += metrics.get("avg_post_reach", 0)

        channels_count = len(telegram_accounts)
        V_avg = aggregated["V"] / aggregated["P"] if aggregated["P"] > 0 else 0
        E_avg = aggregated["E"] / aggregated["P"] if aggregated["P"] > 0 else 0
        ER_view = (
            (aggregated["E"] / aggregated["V"] * 100) if aggregated["V"] > 0 else 0
        )
        ERR_avg = (
            aggregated["err_percent_sum"] / channels_count if channels_count > 0 else 0
        )
        ER_avg = (
            aggregated["er_percent_sum"] / channels_count if channels_count > 0 else 0
        )
        avg_post_reach = (
            aggregated["avg_post_reach_sum"] / channels_count
            if channels_count > 0
            else 0
        )
        ci_index = (
            aggregated["ci_index_sum"] / channels_count if channels_count > 0 else 0
        )

        ws.cell(row=row, column=1, value=author.name)
        ws.cell(row=row, column=2, value=channels_count)
        ws.cell(row=row, column=3, value=aggregated["F"])
        ws.cell(row=row, column=4, value=aggregated["P"])
        ws.cell(row=row, column=5, value=aggregated["V"])
        ws.cell(row=row, column=6, value=round(V_avg, 0))
        ws.cell(row=row, column=7, value=aggregated["E"])
        ws.cell(row=row, column=8, value=round(E_avg, 0))
        ws.cell(row=row, column=9, value=round(ER_view, 2))
        ws.cell(row=row, column=10, value=round(ERR_avg, 2))
        ws.cell(row=row, column=11, value=round(ER_avg, 2))
        ws.cell(row=row, column=12, value=round(avg_post_reach, 0))
        ws.cell(row=row, column=13, value=round(ci_index, 2))
        ws.cell(row=row, column=14, value=aggregated["total_reactions"])
        ws.cell(row=row, column=15, value=aggregated["total_comments"])
        ws.cell(row=row, column=16, value=aggregated["total_shares"])

        row += 1

    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = (
        f"telegram_all_authors_{current_start.date()}_to_{current_end.date()}.xlsx"
    )

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel/{social_account_id}")
async def generate_telegram_excel(
    social_account_id: int,
    current_start: datetime = Query(..., description="Начало текущего периода"),
    current_end: datetime = Query(..., description="Конец текущего периода"),
    previous_start: datetime | None = Query(
        None, description="Начало прошлого периода"
    ),
    previous_end: datetime | None = Query(None, description="Конец прошлого периода"),
):
    """
    Генерирует Excel отчет для Telegram канала

    Возвращает файл Excel с полной аналитикой канала
    """
    # Проверяем существование аккаунта
    social_account = await SocialAccount.filter(id=social_account_id).first()
    if not social_account:
        raise HTTPException(status_code=404, detail="Social account not found")

    if social_account.platform != "telegram":
        raise HTTPException(
            status_code=400, detail="This endpoint only supports Telegram accounts"
        )

    # Вычисляем метрики
    current_metrics = await _calculate_telegram_period_metrics(
        social_account, current_start, current_end
    )

    previous_metrics = None
    comparison_metrics = None

    if previous_start and previous_end:
        previous_metrics = await _calculate_telegram_period_metrics(
            social_account, previous_start, previous_end
        )
        comparison_metrics = _calculate_comparison_metrics(
            current_metrics, previous_metrics
        )

    # Формируем данные для отчета
    report_data = {
        "social_account_id": social_account_id,
        "platform": social_account.platform,
        "username": social_account.username,
        "current_period": {
            "start": current_start.isoformat(),
            "end": current_end.isoformat(),
            "metrics": current_metrics,
        },
        "previous_period": {
            "start": previous_start.isoformat() if previous_start else None,
            "end": previous_end.isoformat() if previous_end else None,
            "metrics": previous_metrics,
        }
        if previous_metrics
        else None,
        "comparison": comparison_metrics,
    }

    # Генерируем Excel
    excel_file = generate_telegram_excel_report(report_data)

    # Формируем имя файла
    filename = f"telegram_report_{social_account.username or social_account_id}_{current_start.date()}_to_{current_end.date()}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

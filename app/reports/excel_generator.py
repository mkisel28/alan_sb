from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, Any
from .utils import get_platform_name


class ExcelReportGenerator:
    """Генерирует отчеты в формате Excel с полными данными"""

    def __init__(self):
        self.wb = None

    def generate(self, data: Dict[str, Any]) -> BytesIO:
        """
        Генерирует Excel отчет со всеми доступными данными

        Args:
            data: Данные аналитики из API

        Returns:
            BytesIO с Excel файлом
        """
        self.wb = Workbook()

        # Удаляем дефолтный лист
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Создаем сводный лист
        self._create_summary_sheet(data)

        # Создаем листы по платформам
        platforms = data.get("platforms", {})
        for platform_key, platform_data in platforms.items():
            self._create_platform_sheet(platform_key, platform_data)

        # Сохраняем в BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        return output

    def _create_summary_sheet(self, data: Dict[str, Any]):
        """Создает сводный лист с агрегированными данными"""
        ws = self.wb.create_sheet("Сводка")

        # Заголовок
        ws["A1"] = "СВОДНЫЙ ОТЧЕТ"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:R1")
        ws.row_dimensions[1].height = 30

        # Период
        period = data.get("period", {})
        ws["A3"] = "Период анализа:"
        ws["B3"] = f"{period.get('start', '')[:10]} — {period.get('end', '')[:10]}"
        ws["A3"].font = Font(bold=True)

        # Заголовки таблицы
        headers = [
            "Платформа",
            "Авторов",
            "Подписчики (F)",
            "Изменение подписчиков",
            "Изменение подписчиков %",
            "Посты (P)",
            "Изменение постов",
            "Изменение постов %",
            "Просмотры (V)",
            "Изменение просмотров",
            "Изменение просмотров %",
            "Вовлечения (E)",
            "Изменение вовлечений",
            "Изменение вовлечений %",
            "Средний Presence Score",
            "Средний ER %",
            "Изменение ER %",
            "Средний Momentum Score",
        ]

        row = 5
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Данные
        platforms = data.get("platforms", {})
        row = 6

        for platform_key, platform_data in platforms.items():
            agg = platform_data["aggregated"]
            deltas = agg.get("deltas", {})
            platform_name = get_platform_name(platform_key)

            ws.cell(row=row, column=1, value=platform_name)
            ws.cell(row=row, column=2, value=agg["total_authors"])
            ws.cell(row=row, column=3, value=agg["total_followers"])
            ws.cell(
                row=row, column=4, value=deltas.get("followers", {}).get("absolute", 0)
            )
            ws.cell(
                row=row, column=5, value=deltas.get("followers", {}).get("percent", 0)
            )
            ws.cell(row=row, column=6, value=agg["total_posts"])
            ws.cell(row=row, column=7, value=deltas.get("posts", {}).get("absolute", 0))
            ws.cell(row=row, column=8, value=deltas.get("posts", {}).get("percent", 0))
            ws.cell(row=row, column=9, value=agg["total_views"])
            ws.cell(
                row=row, column=10, value=deltas.get("views", {}).get("absolute", 0)
            )
            ws.cell(row=row, column=11, value=deltas.get("views", {}).get("percent", 0))
            ws.cell(row=row, column=12, value=agg["total_engagement"])
            ws.cell(
                row=row,
                column=13,
                value=deltas.get("engagement", {}).get("absolute", 0),
            )
            ws.cell(
                row=row, column=14, value=deltas.get("engagement", {}).get("percent", 0)
            )
            ws.cell(row=row, column=15, value=round(agg["avg_PS"], 4))
            ws.cell(row=row, column=16, value=round(agg["avg_ER_view"] * 100, 4))
            ws.cell(
                row=row, column=17, value=deltas.get("ER_view", {}).get("percent", 0)
            )
            ws.cell(row=row, column=18, value=round(agg.get("avg_MS", 0), 4))

            row += 1

        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            if col == 1:  # Платформа
                ws.column_dimensions[get_column_letter(col)].width = 18
            else:
                ws.column_dimensions[get_column_letter(col)].width = 16

        # Увеличиваем высоту строки заголовков
        ws.row_dimensions[5].height = 30

        # Границы
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_cells in ws[f"A5:{get_column_letter(len(headers))}{row - 1}"]:
            for cell in row_cells:
                cell.border = thin_border

        # Форматирование чисел в сводке
        for r in range(6, row):
            # Числовые колонки (авторы, подписчики, посты, просмотры, вовлечения и их дельты)
            for c in [2, 3, 4, 6, 7, 9, 10, 12, 13]:
                cell = ws.cell(row=r, column=c)
                cell.number_format = "#,##0"
            # Проценты
            for c in [5, 8, 11, 14, 15, 16, 17, 18]:
                cell = ws.cell(row=r, column=c)
                cell.number_format = "0.0000"

    def _create_platform_sheet(self, platform_key: str, platform_data: Dict[str, Any]):
        """Создает детальный лист по платформе со всеми метриками"""
        platform_name = get_platform_name(platform_key)
        ws = self.wb.create_sheet(platform_name)

        # Определяем заголовки сразу (нужно для merge_cells)
        headers = [
            # Базовая информация
            "Автор",
            "Username",
            # Текущий период - основные метрики
            "F\nПодписчики",
            "Δ F\nИзменение подписчиков",
            "Δ F %\nИзменение подписчиков %",
            "P\nКол-во постов",
            "V\nОбщие просмотры",
            "V_avg\nСредние просмотры на пост",
            "E\nОбщие вовлечения",
            "E_avg\nСредние вовлечения на пост",
            # Текущий период - метрики вовлеченности
            "ER_view %\nEngagement Rate от просмотров",
            "ER_fol %\nEngagement Rate от подписчиков",
            "SR %\nShare Rate (доля расшариваний)",
            "CR %\nComment Rate (доля комментариев)",
            "Shares\nКол-во расшариваний",
            "Comments\nКол-во комментариев",
            # Предыдущий период - основные метрики
            "prev_F\nПодписчики (пред. период)",
            "prev_P\nКол-во постов (пред. период)",
            "prev_V\nОбщие просмотры (пред. период)",
            "prev_V_avg\nСредние просмотры (пред. период)",
            "prev_E\nОбщие вовлечения (пред. период)",
            "prev_E_avg\nСредние вовлечения (пред. период)",
            # Предыдущий период - метрики вовлеченности
            "prev_ER_view %\nER от просмотров (пред. период)",
            "prev_ER_fol %\nER от подписчиков (пред. период)",
            "prev_SR %\nShare Rate (пред. период)",
            "prev_CR %\nComment Rate (пред. период)",
            "prev_Shares\nРасшаривания (пред. период)",
            "prev_Comments\nКомментарии (пред. период)",
            # Дельты
            "Δ V_avg %\nИзменение средних просмотров %",
            "Δ ER %\nИзменение вовлеченности %",
            # Скоры
            "PS\nPresence Score (уровень присутствия)",
            "prev_PS\nPresence Score (пред. период)",
            "Δ PS\nИзменение PS",
            "MS\nMomentum Score (динамика роста)",
            # Перцентили текущего периода
            "pct_V_avg\nПерцентиль средних просмотров",
            "pct_ER\nПерцентиль вовлеченности",
            "pct_SR\nПерцентиль расшариваний",
            "pct_P\nПерцентиль постов",
            "pct_F\nПерцентиль подписчиков",
            # Перцентили предыдущего периода
            "prev_pct_V_avg\nПерцентиль просмотров (пред.)",
            "prev_pct_ER\nПерцентиль вовлеченности (пред.)",
            "prev_pct_SR\nПерцентиль расшариваний (пред.)",
            "prev_pct_P\nПерцентиль постов (пред.)",
            "prev_pct_F\nПерцентиль подписчиков (пред.)",
            # Перцентили моментума
            "mom_Δ V_avg\nПерцентиль изменения просмотров",
            "mom_Δ ER\nПерцентиль изменения вовлеченности",
            "mom_Δ F\nПерцентиль изменения подписчиков",
        ]

        # Заголовок платформы
        ws["A1"] = platform_name
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws.row_dimensions[1].height = 25

        # Заголовки колонок
        row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Увеличиваем высоту строки заголовков для двухстрочного текста
        ws.row_dimensions[3].height = 60

        # Данные авторов
        row = 4
        authors = platform_data["authors"]

        for author_data in authors:
            metrics = author_data["metrics"]
            scores = author_data["scores"]
            prev_metrics = metrics.get("prev_metrics", {})
            percentiles = scores.get("percentiles", {})
            prev_percentiles = scores.get("prev_percentiles", {})
            momentum_percentiles = scores.get("momentum_percentiles", {})

            col = 1
            # Базовая информация
            ws.cell(row=row, column=col, value=author_data["author_name"])
            col += 1
            ws.cell(row=row, column=col, value=author_data["username"])
            col += 1

            # Текущий период - основные метрики
            ws.cell(row=row, column=col, value=metrics["F"])
            col += 1
            ws.cell(row=row, column=col, value=metrics["delta_F"])
            col += 1
            ws.cell(
                row=row,
                column=col,
                value=round(metrics.get("delta_F_percent", 0) * 100, 4),
            )
            col += 1
            ws.cell(row=row, column=col, value=metrics["P"])
            col += 1
            ws.cell(row=row, column=col, value=metrics["V"])
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["V_avg"], 4))
            col += 1
            ws.cell(row=row, column=col, value=metrics["E"])
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["E_avg"], 4))
            col += 1

            # Текущий период - метрики вовлеченности
            ws.cell(row=row, column=col, value=round(metrics["ER_view"] * 100, 4))
            col += 1
            # ER_fol может быть очень большим если F был 0, ограничим
            er_fol = metrics.get("ER_fol", 0)
            ws.cell(
                row=row, column=col, value=round(er_fol * 100, 4) if er_fol < 100 else 0
            )
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["SR"] * 100, 4))
            col += 1
            ws.cell(row=row, column=col, value=round(metrics["CR"] * 100, 4))
            col += 1
            ws.cell(row=row, column=col, value=metrics.get("total_shares", 0))
            col += 1
            ws.cell(row=row, column=col, value=metrics.get("total_comments", 0))
            col += 1

            # Предыдущий период - основные метрики
            ws.cell(row=row, column=col, value=prev_metrics.get("F", 0))
            col += 1
            ws.cell(row=row, column=col, value=prev_metrics.get("P", 0))
            col += 1
            ws.cell(row=row, column=col, value=prev_metrics.get("V", 0))
            col += 1
            ws.cell(row=row, column=col, value=round(prev_metrics.get("V_avg", 0), 4))
            col += 1
            ws.cell(row=row, column=col, value=prev_metrics.get("E", 0))
            col += 1
            ws.cell(row=row, column=col, value=round(prev_metrics.get("E_avg", 0), 4))
            col += 1

            # Предыдущий период - метрики вовлеченности
            ws.cell(
                row=row,
                column=col,
                value=round(prev_metrics.get("ER_view", 0) * 100, 4),
            )
            col += 1
            prev_er_fol = prev_metrics.get("ER_fol", 0)
            ws.cell(
                row=row,
                column=col,
                value=round(prev_er_fol * 100, 4) if prev_er_fol < 100 else 0,
            )
            col += 1
            ws.cell(
                row=row, column=col, value=round(prev_metrics.get("SR", 0) * 100, 4)
            )
            col += 1
            ws.cell(
                row=row, column=col, value=round(prev_metrics.get("CR", 0) * 100, 4)
            )
            col += 1
            ws.cell(row=row, column=col, value=prev_metrics.get("total_shares", 0))
            col += 1
            ws.cell(row=row, column=col, value=prev_metrics.get("total_comments", 0))
            col += 1

            # Дельты
            ws.cell(
                row=row,
                column=col,
                value=round(metrics.get("delta_V_avg_percent", 0) * 100, 4),
            )
            col += 1
            ws.cell(
                row=row,
                column=col,
                value=round(metrics.get("delta_ER_percent", 0) * 100, 4),
            )
            col += 1

            # Скоры
            ws.cell(row=row, column=col, value=round(scores["PS"], 4))
            col += 1
            ws.cell(row=row, column=col, value=round(scores.get("prev_PS", 0), 4))
            col += 1
            delta_ps = (
                scores["PS"] - scores.get("prev_PS", scores["PS"])
                if "prev_PS" in scores
                else 0
            )
            ws.cell(row=row, column=col, value=round(delta_ps, 4))
            col += 1
            ws.cell(row=row, column=col, value=round(scores.get("MS", 0), 4))
            col += 1

            # Перцентили текущего периода
            ws.cell(row=row, column=col, value=round(percentiles.get("V_avg", 0), 4))
            col += 1
            ws.cell(row=row, column=col, value=round(percentiles.get("ER_view", 0), 4))
            col += 1
            ws.cell(row=row, column=col, value=round(percentiles.get("SR", 0), 4))
            col += 1
            ws.cell(row=row, column=col, value=round(percentiles.get("P", 0), 4))
            col += 1
            ws.cell(row=row, column=col, value=round(percentiles.get("F", 0), 4))
            col += 1

            # Перцентили предыдущего периода
            ws.cell(
                row=row, column=col, value=round(prev_percentiles.get("V_avg", 0), 4)
            )
            col += 1
            ws.cell(
                row=row, column=col, value=round(prev_percentiles.get("ER_view", 0), 4)
            )
            col += 1
            ws.cell(row=row, column=col, value=round(prev_percentiles.get("SR", 0), 4))
            col += 1
            ws.cell(row=row, column=col, value=round(prev_percentiles.get("P", 0), 4))
            col += 1
            ws.cell(row=row, column=col, value=round(prev_percentiles.get("F", 0), 4))
            col += 1

            # Перцентили моментума
            ws.cell(
                row=row,
                column=col,
                value=round(momentum_percentiles.get("delta_V_avg", 0), 4),
            )
            col += 1
            ws.cell(
                row=row,
                column=col,
                value=round(momentum_percentiles.get("delta_ER", 0), 4),
            )
            col += 1
            ws.cell(
                row=row,
                column=col,
                value=round(momentum_percentiles.get("delta_F", 0), 4),
            )
            col += 1

            row += 1

        # Автоширина колонок
        ws.column_dimensions["A"].width = 20  # Автор
        ws.column_dimensions["B"].width = 20  # Username

        # Все остальные колонки - стандартная ширина для метрик
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Границы
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_cells in ws[f"A3:{get_column_letter(len(headers))}{row - 1}"]:
            for cell in row_cells:
                cell.border = thin_border

        # Форматирование чисел
        for r in range(4, row):
            # Целые числа: F, ΔF, P, V, E, Shares, Comments, prev_*
            for c in [3, 4, 6, 7, 9, 15, 16, 17, 18, 19, 21, 27, 28]:
                cell = ws.cell(row=r, column=c)
                cell.number_format = "#,##0"

            # Дробные числа: V_avg, E_avg, prev_V_avg, prev_E_avg
            for c in [8, 10, 20, 22]:
                cell = ws.cell(row=r, column=c)
                cell.number_format = "#,##0.0000"

            # Проценты: все ER, SR, CR, дельты %, перцентили
            for c in [
                5,
                11,
                12,
                13,
                14,
                23,
                24,
                25,
                26,
                29,
                30,
                31,
                32,
                33,
                34,
                35,
                36,
                37,
                38,
                39,
                40,
                41,
                42,
                43,
                44,
                45,
                46,
                47,
                48,
            ]:
                cell = ws.cell(row=r, column=c)
                cell.number_format = "0.0000"

        # Закрепление первой строки и первых двух колонок
        ws.freeze_panes = "C4"

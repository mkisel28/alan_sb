from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from typing import Dict, Any


class TelegramExcelReportGenerator:
    """Генератор Excel отчетов для Telegram каналов"""

    def __init__(self):
        self.wb = None

    def generate(self, data: Dict[str, Any]) -> BytesIO:
        """
        Генерирует Excel отчет для Telegram канала

        Args:
            data: Данные аналитики Telegram канала

        Returns:
            BytesIO с Excel файлом
        """
        self.wb = Workbook()

        # Удаляем дефолтный лист
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Создаем листы
        self._create_summary_sheet(data)
        self._create_metrics_sheet(data)

        # Сохраняем в BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        return output

    def _create_summary_sheet(self, data: Dict[str, Any]):
        """Создает сводный лист"""
        ws = self.wb.create_sheet("Сводка")

        # Заголовок
        ws["A1"] = "ОТЧЕТ ПО TELEGRAM КАНАЛУ"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="0088CC", end_color="0088CC", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:H1")
        ws.row_dimensions[1].height = 30

        # Информация о канале
        current = data.get("current_period", {})
        metrics = current.get("metrics", {})

        ws["A3"] = "Период анализа:"
        ws["B3"] = f"{current.get('start', '')[:10]} — {current.get('end', '')[:10]}"
        ws["A3"].font = Font(bold=True)

        # Основные метрики
        row = 5
        ws[f"A{row}"] = "ОСНОВНЫЕ ПОКАЗАТЕЛИ"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        ws[f"A{row}"].fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )
        ws.merge_cells(f"A{row}:D{row}")

        row += 2
        headers = [
            ("Подписчики (F)", metrics.get("F", 0)),
            ("Изменение подписчиков (ΔF)", metrics.get("delta_F", 0)),
            ("Изменение подписчиков (%)", f"{metrics.get('delta_F_percent', 0)}%"),
            ("Публикации (P)", metrics.get("P", 0)),
            ("Просмотры (V)", metrics.get("V", 0)),
            ("Средние просмотры (V_avg)", round(metrics.get("V_avg", 0), 2)),
        ]

        for label, value in headers:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Специфичные метрики Telegram
        row += 1
        ws[f"A{row}"] = "МЕТРИКИ TELEGRAM"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        ws[f"A{row}"].fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )
        ws.merge_cells(f"A{row}:D{row}")

        row += 2
        telegram_metrics = [
            ("Средний охват поста", metrics.get("avg_post_reach", 0)),
            ("Рекламный охват (12ч)", metrics.get("adv_post_reach_12h", 0)),
            ("Рекламный охват (24ч)", metrics.get("adv_post_reach_24h", 0)),
            ("Рекламный охват (48ч)", metrics.get("adv_post_reach_48h", 0)),
            ("ERR% (вовлеч. в просмотры)", f"{metrics.get('err_percent', 0)}%"),
            ("ERR24% (вовлеч. 24ч)", f"{metrics.get('err24_percent', 0)}%"),
            ("ER% (вовлеч. во взаимод.)", f"{metrics.get('er_percent', 0)}%"),
            ("Индекс цитирования (ИЦ)", metrics.get("ci_index", 0)),
            ("Дневной охват", metrics.get("daily_reach", 0)),
            ("Упоминания", metrics.get("mentions_count", 0)),
            ("Пересылки", metrics.get("forwards_count", 0)),
            ("Упоминающих каналов", metrics.get("mentioning_channels_count", 0)),
        ]

        for label, value in telegram_metrics:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Вовлечённость
        row += 1
        ws[f"A{row}"] = "ВОВЛЕЧЁННОСТЬ"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        ws[f"A{row}"].fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )
        ws.merge_cells(f"A{row}:D{row}")

        row += 2
        engagement_metrics = [
            ("Всего взаимодействий (E)", metrics.get("E", 0)),
            ("Реакции", metrics.get("total_reactions", 0)),
            ("Комментарии", metrics.get("total_comments", 0)),
            ("Пересылки", metrics.get("total_shares", 0)),
            ("Среднее на пост (E_avg)", round(metrics.get("E_avg", 0), 2)),
            ("Реакций на пост", round(metrics.get("reactions_per_post", 0), 2)),
            ("Комментариев на пост", round(metrics.get("comments_per_post", 0), 2)),
            ("Пересылок на пост", round(metrics.get("shares_per_post", 0), 2)),
            ("ER_view (%)", f"{metrics.get('ER_view', 0)}%"),
        ]

        for label, value in engagement_metrics:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Устанавливаем ширину колонок
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

    def _create_metrics_sheet(self, data: Dict[str, Any]):
        """Создает лист с детальными метриками"""
        ws = self.wb.create_sheet("Детальные метрики")

        current = data.get("current_period", {})
        previous = data.get("previous_period", {})
        comparison = data.get("comparison", {})

        # Заголовок
        ws["A1"] = "СРАВНЕНИЕ ПЕРИОДОВ"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="0088CC", end_color="0088CC", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:E1")

        # Заголовки колонок
        headers = [
            "Метрика",
            "Текущий период",
            "Прошлый период",
            "Изменение",
            "Изменение (%)",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D0D0D0", end_color="D0D0D0", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Метрики для сравнения
        metrics_to_show = [
            ("Подписчики (F)", "F"),
            ("Публикации (P)", "P"),
            ("Просмотры (V)", "V"),
            ("Средние просмотры", "V_avg"),
            ("Вовлечённость (E)", "E"),
            ("Средняя вовлечённость", "E_avg"),
            ("ER_view (%)", "ER_view"),
            ("ERR%", "ERR_percent"),
            ("ER%", "ER_percent"),
            ("Реакции", "total_reactions"),
            ("Комментарии", "total_comments"),
            ("Пересылки", "total_shares"),
            ("Средний охват", "avg_post_reach"),
            ("Индекс цитирования", "ci_index"),
        ]

        current_metrics = current.get("metrics", {})
        previous_metrics = previous.get("metrics", {}) if previous else {}

        row = 4
        for label, key in metrics_to_show:
            curr_val = current_metrics.get(key, 0) or 0
            prev_val = previous_metrics.get(key, 0) or 0 if previous else "-"

            if comparison and previous:
                delta = comparison.get(f"{key}_delta", 0)
                delta_percent = comparison.get(f"{key}_delta_percent", 0)
            else:
                delta = "-"
                delta_percent = "-"

            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=curr_val)
            ws.cell(row=row, column=3, value=prev_val)
            ws.cell(row=row, column=4, value=delta)
            ws.cell(
                row=row,
                column=5,
                value=f"{delta_percent}%" if delta_percent != "-" else "-",
            )

            # Раскраска изменений
            if delta != "-" and isinstance(delta, (int, float)):
                color = "90EE90" if delta > 0 else "FFB6C1" if delta < 0 else "FFFFFF"
                ws.cell(row=row, column=4).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                ws.cell(row=row, column=5).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

            row += 1

        # Устанавливаем ширину колонок
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15


def generate_telegram_excel_report(data: Dict[str, Any]) -> BytesIO:
    """
    Генерирует Excel отчет для Telegram канала

    Args:
        data: Данные аналитики из API telegram-analytics

    Returns:
        BytesIO с Excel файлом
    """
    generator = TelegramExcelReportGenerator()
    return generator.generate(data)
